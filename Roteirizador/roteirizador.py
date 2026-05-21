#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Roteirizador — otimizador de rotas diarias com janelas de disponibilidade.

Resolve um problema de roteirizacao multi-dia com janelas de tempo (VRPTW):
- agrupa pontos geograficamente proximos no mesmo dia (menor deslocamento);
- respeita os dias da semana em que cada ponto pode ser atendido;
- respeita a janela de horario de cada ponto e a jornada de trabalho do dia;
- ordena as visitas de cada dia para minimizar a distancia total.

Usa SOMENTE a biblioteca padrao do Python (funciona sem pip install).
Le planilhas .csv nativamente; .xlsx so se 'openpyxl' estiver instalado.

Uso basico:
    python3 roteirizador.py --arquivo pontos.csv

Veja --help para todos os parametros.
"""

import argparse
import csv
import json
import math
import os
import re
import sys
import unicodedata
from datetime import date, datetime, timedelta

# ---------------------------------------------------------------------------
# Utilidades de texto
# ---------------------------------------------------------------------------

def normalizar(texto):
    """Minusculas, sem acento, sem espacos nas pontas."""
    if texto is None:
        return ""
    t = unicodedata.normalize("NFKD", str(texto))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return t.strip().lower()


DIAS_PT = ["Segunda", "Terca", "Quarta", "Quinta", "Sexta", "Sabado", "Domingo"]

_MAPA_DIAS = {
    "segunda": 0, "seg": 0, "2a": 0, "monday": 0, "mon": 0,
    "terca": 1, "ter": 1, "3a": 1, "tuesday": 1, "tue": 1,
    "quarta": 2, "qua": 2, "4a": 2, "wednesday": 2, "wed": 2,
    "quinta": 3, "qui": 3, "5a": 3, "thursday": 3, "thu": 3,
    "sexta": 4, "sex": 4, "6a": 4, "friday": 4, "fri": 4,
    "sabado": 5, "sab": 5, "saturday": 5, "sat": 5,
    "domingo": 6, "dom": 6, "sunday": 6, "sun": 6,
}


def parse_dias(texto):
    """Converte texto livre em conjunto de indices de dia da semana (0=Seg..6=Dom)."""
    n = normalizar(texto)
    if not n or n in ("todos", "qualquer", "todos os dias", "qualquer dia", "*"):
        return {0, 1, 2, 3, 4, 5, 6}
    if "util" in n or "uteis" in n or "comercial" in n or "semana" in n and "fim" not in n:
        return {0, 1, 2, 3, 4}
    if "fim de semana" in n or "fds" in n:
        return {5, 6}
    dias = set()
    # tokeniza por separadores comuns
    for token in re.split(r"[,;/|e\s]+", n):
        token = token.strip()
        if not token:
            continue
        # remove sufixo "-feira"
        token = token.replace("-feira", "").replace("feira", "")
        if token in _MAPA_DIAS:
            dias.add(_MAPA_DIAS[token])
        else:
            # tenta prefixo de 3 letras
            chave = token[:3]
            if chave in _MAPA_DIAS:
                dias.add(_MAPA_DIAS[chave])
    return dias if dias else {0, 1, 2, 3, 4, 5, 6}


def _hhmm_para_min(h, m):
    return int(h) * 60 + int(m or 0)


def parse_janela(texto, jornada_padrao):
    """Converte texto livre em (inicio_min, fim_min). jornada_padrao = (ini, fim)."""
    n = normalizar(texto)
    if not n or n in ("qualquer", "integral", "qualquer horario", "livre"):
        return jornada_padrao
    if "manha" in n:
        return (_hhmm_para_min(8, 0), _hhmm_para_min(12, 0))
    if "tarde" in n:
        return (_hhmm_para_min(13, 0), _hhmm_para_min(18, 0))
    if "noite" in n:
        return (_hhmm_para_min(18, 0), _hhmm_para_min(22, 0))
    if "comercial" in n:
        return (_hhmm_para_min(8, 0), _hhmm_para_min(18, 0))
    # procura horarios no formato 8h, 8:00, 08h30
    achados = re.findall(r"(\d{1,2})\s*[:h]\s*(\d{2})?", n)
    if len(achados) >= 2:
        ini = _hhmm_para_min(achados[0][0], achados[0][1])
        fim = _hhmm_para_min(achados[1][0], achados[1][1])
        return (ini, fim) if fim > ini else jornada_padrao
    # fallback: dois numeros separados por as/a/ate/-/–
    if re.search(r"\d.*(?:as|a|ate|-|–| as ).*\d", n):
        nums = re.findall(r"\d{1,2}", n)
        if len(nums) >= 2:
            ini = _hhmm_para_min(nums[0], 0)
            fim = _hhmm_para_min(nums[1], 0)
            return (ini, fim) if fim > ini else jornada_padrao
    return jornada_padrao


def min_para_hhmm(m):
    m = int(round(m))
    return "{:02d}:{:02d}".format(m // 60, m % 60)


def parse_hhmm_arg(texto):
    achados = re.findall(r"(\d{1,2})\s*[:h]\s*(\d{2})?", normalizar(texto))
    if achados:
        return _hhmm_para_min(achados[0][0], achados[0][1])
    raise ValueError("Horario invalido: {}".format(texto))


# ---------------------------------------------------------------------------
# Geografia
# ---------------------------------------------------------------------------

def haversine_km(a, b):
    """Distancia em linha reta (km) entre (lat, lon) a e b."""
    lat1, lon1 = a
    lat2, lon2 = b
    R = 6371.0
    p1 = math.radians(lat1)
    p2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlmb = math.radians(lon2 - lon1)
    h = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlmb / 2) ** 2
    return 2 * R * math.asin(math.sqrt(h))


_CACHE_GEO = None
_CACHE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".cache_geocode.json")


def _carregar_cache():
    global _CACHE_GEO
    if _CACHE_GEO is None:
        try:
            with open(_CACHE_PATH, "r", encoding="utf-8") as f:
                _CACHE_GEO = json.load(f)
        except Exception:
            _CACHE_GEO = {}
    return _CACHE_GEO


def _salvar_cache():
    try:
        with open(_CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump(_CACHE_GEO, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def geocodificar(endereco):
    """Retorna (lat, lon) via Nominatim/OpenStreetMap, com cache. None se falhar."""
    import time
    from urllib.parse import urlencode
    from urllib.request import Request, urlopen

    cache = _carregar_cache()
    chave = normalizar(endereco)
    if chave in cache and cache[chave]:
        return tuple(cache[chave])
    try:
        url = "https://nominatim.openstreetmap.org/search?" + urlencode(
            {"q": endereco, "format": "json", "limit": 1, "countrycodes": "br"}
        )
        req = Request(url, headers={"User-Agent": "Roteirizador/1.0 (planejamento de rotas)"})
        with urlopen(req, timeout=15) as resp:
            dados = json.load(resp)
        time.sleep(1.1)  # politica de uso do Nominatim: <=1 req/s
        if dados:
            coord = (float(dados[0]["lat"]), float(dados[0]["lon"]))
            cache[chave] = list(coord)
            _salvar_cache()
            return coord
    except Exception as e:
        print("  ! falha ao geocodificar '{}': {}".format(endereco, e), file=sys.stderr)
    cache[chave] = None
    _salvar_cache()
    return None


# ---------------------------------------------------------------------------
# Leitura da planilha
# ---------------------------------------------------------------------------

# nomes de coluna aceitos (normalizados) -> campo interno
_ALIASES = {
    "nome": ["nome", "ponto", "cliente", "local", "id", "identificacao", "estabelecimento"],
    "endereco": ["endereco", "address", "logradouro", "local completo", "endereco completo"],
    "lat": ["lat", "latitude"],
    "lon": ["lon", "lng", "long", "longitude"],
    "dias": ["dias", "dias disponiveis", "disponibilidade", "disponibilidade dias",
             "dias da semana", "dia"],
    "janela": ["janela", "horario", "horarios", "horario disponivel", "disponibilidade horario",
               "horario de atendimento", "horario disponivel"],
    "duracao": ["duracao", "tempo", "tempo atendimento", "duracao visita", "minutos"],
    "prioridade": ["prioridade", "frequencia"],
}


def _mapear_colunas(cabecalho):
    mapa = {}
    norm = {normalizar(c): c for c in cabecalho}
    for campo, nomes in _ALIASES.items():
        for nome in nomes:
            if nome in norm:
                mapa[campo] = norm[nome]
                break
    return mapa


def _linhas_do_arquivo(caminho):
    ext = os.path.splitext(caminho)[1].lower()
    if ext in (".csv", ".txt", ""):
        # detecta separador
        with open(caminho, "r", encoding="utf-8-sig", newline="") as f:
            amostra = f.read(4096)
            f.seek(0)
            try:
                dialeto = csv.Sniffer().sniff(amostra, delimiters=",;\t")
            except Exception:
                dialeto = csv.excel
            leitor = csv.reader(f, dialeto)
            linhas = [list(r) for r in leitor]
        return linhas
    if ext in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("ERRO: para ler .xlsx instale openpyxl (pip install openpyxl) "
                     "ou exporte a planilha como .csv.")
        wb = load_workbook(caminho, read_only=True, data_only=True)
        ws = wb.active
        return [[("" if c is None else c) for c in row] for row in ws.iter_rows(values_only=True)]
    sys.exit("ERRO: formato nao suportado: {}. Use .csv ou .xlsx.".format(ext))


def ler_pontos(caminho, jornada_padrao, duracao_padrao, geocodificar_faltantes):
    linhas = [l for l in _linhas_do_arquivo(caminho) if any(str(c).strip() for c in l)]
    if len(linhas) < 2:
        sys.exit("ERRO: planilha vazia ou sem dados.")
    cabecalho = [str(c) for c in linhas[0]]
    mapa = _mapear_colunas(cabecalho)
    if "nome" not in mapa:
        mapa["nome"] = cabecalho[0]
    tem_coords = "lat" in mapa and "lon" in mapa
    if not tem_coords and "endereco" not in mapa:
        sys.exit("ERRO: a planilha precisa de colunas de latitude/longitude OU de endereco.\n"
                 "Colunas encontradas: {}".format(", ".join(cabecalho)))

    idx = {campo: cabecalho.index(col) for campo, col in mapa.items()}
    pontos = []
    sem_geo = []
    for n, linha in enumerate(linhas[1:], start=2):
        def val(campo):
            i = idx.get(campo)
            return str(linha[i]).strip() if i is not None and i < len(linha) else ""

        nome = val("nome") or "Ponto {}".format(n)
        coord = None
        if tem_coords and val("lat") and val("lon"):
            try:
                coord = (float(val("lat").replace(",", ".")), float(val("lon").replace(",", ".")))
            except ValueError:
                coord = None
        endereco = val("endereco")
        if coord is None and endereco and geocodificar_faltantes:
            print("  geocodificando: {} ...".format(nome), file=sys.stderr)
            coord = geocodificar(endereco)
        if coord is None:
            sem_geo.append(nome)
            continue

        dur = duracao_padrao
        if val("duracao"):
            try:
                dur = float(re.findall(r"\d+", val("duracao"))[0])
            except (IndexError, ValueError):
                pass
        prioridade = 0
        if val("prioridade"):
            try:
                prioridade = float(re.findall(r"\d+", val("prioridade"))[0])
            except (IndexError, ValueError):
                pass

        pontos.append({
            "nome": nome,
            "endereco": endereco,
            "coord": coord,
            "dias": parse_dias(val("dias")),
            "janela": parse_janela(val("janela"), jornada_padrao),
            "duracao": dur,
            "prioridade": prioridade,
        })
    return pontos, sem_geo


# ---------------------------------------------------------------------------
# Roteirizacao (VRPTW heuristico)
# ---------------------------------------------------------------------------

class Cfg:
    pass


def simular(rota, cfg):
    """Simula a jornada de um dia. Retorna dict com viabilidade, chegadas, km, fim."""
    cur = cfg.inicio
    loc = cfg.origem  # pode ser None (sem deposito): comeca no 1o ponto
    total_km = 0.0
    chegadas = []
    espera = 0.0
    primeira = True
    for pt in rota:
        if loc is not None:
            d = haversine_km(loc, pt["coord"])
            total_km += d
            cur += (d / cfg.velocidade) * 60.0
        elif not primeira:
            d = haversine_km(rota[0]["coord"], pt["coord"])  # nunca ocorre; guarda
        win_i, win_f = pt["janela"]
        if cur < win_i:
            espera += win_i - cur
            cur = win_i
        if cur > win_f + 1e-6:
            return {"viavel": False}
        chegadas.append(cur)
        cur += pt["duracao"]
        if cur > cfg.fim + 1e-6:
            return {"viavel": False}
        loc = pt["coord"]
        primeira = False
    if cfg.origem is not None and rota:
        total_km += haversine_km(loc, cfg.origem)
    return {"viavel": True, "chegadas": chegadas, "km": total_km, "fim": cur, "espera": espera}


def viavel(rota, cfg):
    if cfg.capacidade and len(rota) > cfg.capacidade:
        return False
    return simular(rota, cfg)["viavel"]


def custo_km(rota, cfg):
    r = simular(rota, cfg)
    return r["km"] if r["viavel"] else float("inf")


def melhor_insercao(rota, pt, cfg):
    """Posicao de menor custo para inserir pt na rota. Retorna (custo_extra, pos) ou None."""
    base = custo_km(rota, cfg)
    if base == float("inf"):
        base = 0.0
    melhor = None
    for pos in range(len(rota) + 1):
        nova = rota[:pos] + [pt] + rota[pos:]
        if not viavel(nova, cfg):
            continue
        extra = custo_km(nova, cfg) - base
        if melhor is None or extra < melhor[0]:
            melhor = (extra, pos)
    return melhor


def dois_opt(rota, cfg):
    """Melhora a ordem da rota com 2-opt mantendo viabilidade."""
    if len(rota) < 4:
        return rota
    melhorou = True
    melhor = rota
    melhor_custo = custo_km(melhor, cfg)
    while melhorou:
        melhorou = False
        for i in range(len(melhor) - 1):
            for j in range(i + 1, len(melhor)):
                nova = melhor[:i] + melhor[i:j + 1][::-1] + melhor[j + 1:]
                if not viavel(nova, cfg):
                    continue
                c = custo_km(nova, cfg)
                if c < melhor_custo - 1e-9:
                    melhor, melhor_custo = nova, c
                    melhorou = True
    return melhor


def gerar_slots(cfg, n_pontos):
    """Gera datas de atendimento (apenas dias uteis configurados), o suficiente p/ os pontos."""
    slots = []
    d = cfg.data_inicio
    limite = cfg.data_inicio + timedelta(days=120)
    # capacidade teorica por dia limita o numero de slots necessarios
    cap = cfg.capacidade or 999
    max_slots = max(1, math.ceil(n_pontos / cap)) + cfg.dias_extras
    while d <= limite and len(slots) < max_slots:
        if d.weekday() in cfg.dias_uteis:
            slots.append(d)
        d += timedelta(days=1)
    return slots


def roteirizar(pontos, cfg):
    slots = gerar_slots(cfg, len(pontos))
    rotas = {s: [] for s in slots}
    nao_alocados = []

    # mais restritos primeiro: menos dias disponiveis, janela mais curta, fecha mais cedo
    def restricao(p):
        dias_validos = len([s for s in slots if s.weekday() in p["dias"]])
        win_i, win_f = p["janela"]
        return (dias_validos, win_f - win_i, win_f, -p["prioridade"])

    ordem = sorted(pontos, key=restricao)

    indice = {s: i for i, s in enumerate(slots)}
    for pt in ordem:
        candidatos = []
        for s in slots:
            if s.weekday() not in pt["dias"]:
                continue
            ins = melhor_insercao(rotas[s], pt, cfg)
            if ins is not None:
                extra, pos = ins
                # penaliza abrir um dia novo: preferimos empacotar pontos
                # proximos no mesmo dia ate esgotar jornada/capacidade.
                novo = len(rotas[s]) == 0
                score = extra + (cfg.penalidade_novo_dia if novo else 0.0)
                candidatos.append((score, indice[s], s, pos))
        if not candidatos:
            motivo = "sem dia/horario viavel na jornada" if any(
                s.weekday() in pt["dias"] for s in slots) else "nenhum slot no dia permitido"
            nao_alocados.append((pt, motivo))
            continue
        # menor custo; empata pelo slot mais cedo (preenche os primeiros dias)
        candidatos.sort(key=lambda c: (round(c[0], 3), c[1]))
        _, _, s, pos = candidatos[0]
        rotas[s].insert(pos, pt)

    # refino local por dia
    for s in slots:
        if len(rotas[s]) >= 4:
            rotas[s] = dois_opt(rotas[s], cfg)

    return slots, rotas, nao_alocados


# ---------------------------------------------------------------------------
# Saida
# ---------------------------------------------------------------------------

def link_maps(rota, cfg):
    pts = []
    if cfg.origem is not None:
        pts.append(cfg.origem)
    pts += [p["coord"] for p in rota]
    if cfg.origem is not None:
        pts.append(cfg.origem)
    if len(pts) < 2:
        return ""
    return "https://www.google.com/maps/dir/" + "/".join(
        "{:.6f},{:.6f}".format(lat, lon) for lat, lon in pts)


def imprimir(slots, rotas, nao_alocados, cfg):
    total_km_geral = 0.0
    total_pontos = 0
    print("\n" + "=" * 64)
    print("  ROTEIRIZACAO DIARIA — plano otimizado")
    print("  Jornada {}-{} | velocidade {} km/h | {}".format(
        min_para_hhmm(cfg.inicio), min_para_hhmm(cfg.fim), cfg.velocidade,
        "deposito definido" if cfg.origem else "sem ponto de partida fixo"))
    print("=" * 64)

    for s in slots:
        rota = rotas[s]
        if not rota:
            continue
        sim = simular(rota, cfg)
        total_km_geral += sim["km"]
        total_pontos += len(rota)
        print("\n### {} — {}  ({} pontos | {:.1f} km | termina {})".format(
            DIAS_PT[s.weekday()], s.strftime("%d/%m/%Y"), len(rota),
            sim["km"], min_para_hhmm(sim["fim"])))
        print("    {:<5} {:<28} {:<13} {:<13}".format("#", "Ponto", "Chegada", "Janela"))
        for i, (pt, cheg) in enumerate(zip(rota, sim["chegadas"]), 1):
            wi, wf = pt["janela"]
            print("    {:<5} {:<28} {:<13} {:<13}".format(
                i, pt["nome"][:28], min_para_hhmm(cheg),
                "{}-{}".format(min_para_hhmm(wi), min_para_hhmm(wf))))
        print("    Mapa: {}".format(link_maps(rota, cfg)))

    print("\n" + "-" * 64)
    print("  TOTAL: {} pontos atendidos em {} dias | {:.1f} km".format(
        total_pontos, sum(1 for s in slots if rotas[s]), total_km_geral))
    if nao_alocados:
        print("\n  NAO ALOCADOS ({}):".format(len(nao_alocados)))
        for pt, motivo in nao_alocados:
            print("    - {} ({})".format(pt["nome"], motivo))
        print("  -> aumente o horizonte (--dias-extras), a jornada, a capacidade,")
        print("     ou verifique as janelas/dias destes pontos.")
    print("=" * 64 + "\n")


def exportar_csv(caminho, slots, rotas, cfg):
    with open(caminho, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["data", "dia_semana", "ordem", "ponto", "endereco",
                    "chegada", "janela_inicio", "janela_fim", "duracao_min", "lat", "lon"])
        for s in slots:
            rota = rotas[s]
            if not rota:
                continue
            sim = simular(rota, cfg)
            for i, (pt, cheg) in enumerate(zip(rota, sim["chegadas"]), 1):
                wi, wf = pt["janela"]
                w.writerow([s.strftime("%Y-%m-%d"), DIAS_PT[s.weekday()], i, pt["nome"],
                            pt["endereco"], min_para_hhmm(cheg), min_para_hhmm(wi),
                            min_para_hhmm(wf), int(pt["duracao"]),
                            "{:.6f}".format(pt["coord"][0]), "{:.6f}".format(pt["coord"][1])])
    print("CSV salvo em: {}".format(caminho))


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv=None):
    p = argparse.ArgumentParser(
        description="Otimizador de rotas diarias com janelas de disponibilidade.")
    p.add_argument("--arquivo", required=True, help="planilha de pontos (.csv ou .xlsx)")
    p.add_argument("--inicio", default="08:00", help="inicio da jornada (default 08:00)")
    p.add_argument("--fim", default="18:00", help="fim da jornada (default 18:00)")
    p.add_argument("--velocidade", type=float, default=30.0, help="km/h medio (default 30)")
    p.add_argument("--duracao-padrao", type=float, default=30.0,
                   help="minutos por visita quando nao informado (default 30)")
    p.add_argument("--capacidade", type=int, default=0,
                   help="max de pontos por dia (0 = limitado so pela jornada)")
    p.add_argument("--dias-uteis", default="seg,ter,qua,qui,sex",
                   help="dias permitidos para rotas (default seg-sex)")
    p.add_argument("--data-inicio", default="",
                   help="data inicial AAAA-MM-DD (default: hoje)")
    p.add_argument("--dias-extras", type=int, default=2,
                   help="slots extras alem do minimo teorico (default 2)")
    p.add_argument("--origem", default="",
                   help="deposito/partida 'lat,lon' (opcional)")
    p.add_argument("--sem-geocodificar", action="store_true",
                   help="nao tenta geocodificar enderecos (exige lat/lon na planilha)")
    p.add_argument("--saida", default="", help="caminho do CSV de saida (opcional)")
    args = p.parse_args(argv)

    cfg = Cfg()
    cfg.inicio = parse_hhmm_arg(args.inicio)
    cfg.fim = parse_hhmm_arg(args.fim)
    cfg.velocidade = args.velocidade
    cfg.capacidade = args.capacidade or 0
    cfg.dias_uteis = parse_dias(args.dias_uteis)
    cfg.dias_extras = max(0, args.dias_extras)
    cfg.penalidade_novo_dia = 100000.0
    cfg.data_inicio = (datetime.strptime(args.data_inicio, "%Y-%m-%d").date()
                       if args.data_inicio else date.today())
    cfg.origem = None
    if args.origem:
        try:
            lat, lon = [float(x) for x in args.origem.replace(" ", "").split(",")]
            cfg.origem = (lat, lon)
        except Exception:
            sys.exit("ERRO: --origem deve ser 'lat,lon'.")

    jornada = (cfg.inicio, cfg.fim)
    pontos, sem_geo = ler_pontos(args.arquivo, jornada, args.duracao_padrao,
                                 not args.sem_geocodificar)
    if sem_geo:
        print("AVISO: {} ponto(s) sem coordenada (ignorados): {}".format(
            len(sem_geo), ", ".join(sem_geo)), file=sys.stderr)
    if not pontos:
        sys.exit("ERRO: nenhum ponto com coordenada valida para roteirizar.")

    slots, rotas, nao_alocados = roteirizar(pontos, cfg)
    imprimir(slots, rotas, nao_alocados, cfg)
    if args.saida:
        exportar_csv(args.saida, slots, rotas, cfg)


if __name__ == "__main__":
    main()
